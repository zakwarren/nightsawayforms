import openpyxl
import datetime
import json


def equipment(wb, ws, bookName, directory, leader, myGroup, event):
    """Build the equipment request form for the camp"""
    print("Generating equipment request form...")

    # generate border styles
    upBorder = openpyxl.styles.borders.Border(top = openpyxl.styles.borders.Side(style='thin'))
    downBorder = openpyxl.styles.borders.Border(bottom = openpyxl.styles.borders.Side(style='thin'))
    leftBorder = openpyxl.styles.borders.Border(left = openpyxl.styles.borders.Side(style='thin'))
    # rightBorder = openpyxl.styles.borders.Border(right = openpyxl.styles.borders.Side(style='thin'))

    # create text wrap style
    wrapText = openpyxl.styles.Alignment(wrap_text=True)
    # create date format syle
    dateStyle = openpyxl.styles.NamedStyle(name='datetime', number_format='DD/MM/YYYY')

    # apply covering box borders
    for rows in range(8, 19):
        ws.cell(column=12, row=rows).border = leftBorder
        ws.cell(column=1, row=rows).border = leftBorder

    for col in range(1, 12):
        ws.cell(column=col, row=8).border = upBorder
        ws.cell(column=col, row=18).border = downBorder
        ws.cell(column=col, row=19).border = downBorder

    # add covering details
    ws.cell(column=2, row=9).value = "Colony / Pack / Troop" # B9
    ws.cell(column=5, row=9).value = myGroup.section # E9

    ws.cell(column=2, row=11).value = "Request form submission date" # B11
    ws.cell(column=6, row=11).value = datetime.date.today() # F11
    ws.cell(column=6, row=11).style = dateStyle

    ws.cell(column=2, row=13).value = "Equipment required date" # B13
    ws.cell(column=6, row=13).value = event.startDate # F13
    ws.cell(column=6, row=13).style = dateStyle

    ws.cell(column=2, row=15).value = "Equipment return date" # B15
    ws.cell(column=6, row=15).value = event.endDate # F15
    ws.cell(column=6, row=15).style = dateStyle

    ws.cell(column=2, row=17).value = "Signed" # B17
    ws.cell(column=5, row=17).value = "Name" # E17
    ws.cell(column=6, row=17).value = leader.name # F17
    ws.cell(column=8, row=17).value = "Position" # G17
    ws.cell(column=9, row=17).value = leader.position # H17

    ws.cell(column=8, row=13).value = "Time" # H13
    ws.cell(column=9, row=13).value = "18:00" # I13
    ws.cell(column=8, row=15).value = "Time" # H15
    ws.cell(column=9, row=15).value = "16:00" # I15

    # add covering details underlines
    i = 0
    rows = 9
    while i < 5:
        col = 5
        while rows < 16 and col < 7:
            # underline col E to F
            ws.cell(column=col, row=rows).border = downBorder
            col += 1
        col += 2
        while rows > 12 and col > 8 and col < 11:
            # underline col I to J
            ws.cell(column=col, row=rows).border = downBorder
            col += 1
        rows += 2
        i += 1
    col = 3
    while col < 11:
        # underline C17-D17, F17-G17, I17-J17
        if col != 5 and col != 8:
            ws.cell(column=col, row=17).border = downBorder
        col += 1

    # add table headers
    ws.cell(column=2, row=19).value = "Equipment description"
    ws.cell(column=8, row=19).value = "Equipment identification"
    ws.cell(column=8, row=19).alignment = wrapText
    ws.column_dimensions['H'].width = 13
    ws.cell(column=9, row=19).value = "Quantity requested"
    ws.cell(column=9, row=19).alignment = wrapText
    ws.column_dimensions['I'].width = 13
    ws.cell(column=10, row=19).value = "Quantity supplied"
    ws.cell(column=10, row=19).alignment = wrapText
    ws.column_dimensions['J'].width = 13
    ws.cell(column=11, row=19).value = "Quantity returned"
    ws.cell(column=11, row=19).alignment = wrapText
    ws.column_dimensions['K'].width = 13

    # apply missing table header side borders
    col = 8
    while col < 13:
        ws.cell(column=col, row=19).border = leftBorder
        col += 1

    # pull in equipment list from JSON data
    with open('config/equipment.json') as jsonData:
        allEquip = json.load(jsonData)
    lists = allEquip['equipment']
    i = 20

    # set list of categories with calculated responses
    preDefined = [
        "sleeping",
        "hygiene",
        "light"
    ]

    for key, values in lists.items():
        # add predefined categories as applicable
        if event.nightsAway > 0 and key in preDefined:
            include = 'y'
        
        # manage special categories
        elif key == "catering" and event.catering:
            include = 'y'
        elif key == "sharps" and event.sharps:
            include = 'y'
        elif key == "water activities" and event.waterActivities:
            include = 'y'
        
        # check whether to add each specific section
        else:
            include = input("Include " + key + " section (y/n)? ")
        
        # for each section to include, check through equipment and amount
        if include == 'y':
            for equip in values:
                check = input(key.capitalize() + ": Take " + equip + " (y/n)? ")
                if check == 'y':
                    ws.cell(column=2, row=i).value = equip
                    amount = input("How many? ")
                    ws.cell(column=9, row=i).value = amount
                    # add border to row
                    for col in range(1, 12):
                        ws.cell(column=col, row=i).border = upBorder
                    ws.cell(column=12, row=i).border = leftBorder
                    i += 1
    
    # adds table border to final row
    i += 1
    for col in range(1, 12):
        ws.cell(column=col, row=i).border = upBorder
    
    # add closer
    i += 1
    ws.cell(column=1, row=i).value = "Equipment above has been returned to the store in a clean and dry condition. All defects have been reported."

    # save workbook
    try:
        bookName = 'Equipment request' + bookName
        print("Saving:", bookName)
        wb.save(directory + bookName)
    except:
        print("Failed to save " + bookName)

